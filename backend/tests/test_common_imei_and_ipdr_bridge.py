"""Common IMEI report + IPDR /64 enrichment + target-only CDR IMEI."""

from datetime import datetime, timedelta, timezone

from backend.app.entity_resolution import common_imei
from backend.app.entity_resolution import service as er
from backend.app.normalization import service as norm

IST = timezone(timedelta(hours=5, minutes=30))


def test_common_imei_report_links_phone_to_handset(tmp_path):
    # LEA shape: Number=IMEI, phone columns with Yes
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Number", "Count", "9876543210", "9123456789", "Handset Details"])
    ws.append(["356938035643809", "2", "Yes", "", "Samsung"])
    path = tmp_path / "CDR__1__Common_IMEI_Report.xlsx"
    wb.save(path)

    links = common_imei.load_common_imei_links(str(tmp_path), events=[])
    assert len(links) == 1
    ids = set(links[0]["own_identifiers"])
    assert ("IMEI", "356938035643809") in ids
    assert ("PHONE", "+919876543210") in ids


def test_common_imei_ipdr_session_column_resolves_via_events(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    stem = "16371450_0_25156076_IPV6_IPDR_2409_40d2_132d_b3f8_20241125132132"
    ws.append(["Number", "Count", stem, "Handset Details"])
    ws.append(["35533017092057", "1", "Yes", ""])
    path = tmp_path / "ipdr__1365__IPDR_-_Common_IMEI_Report.xlsx"
    wb.save(path)

    events = [{
        "event_type": "IP_SESSION",
        "primary": ("PHONE", "+917500107305"),
        "own_identifiers": [("PHONE", "+917500107305")],
        "provenance": {"source_file": f"{stem}.csv"},
    }]
    links = common_imei.load_common_imei_links(str(tmp_path), events=events)
    assert len(links) == 1
    ids = set(links[0]["own_identifiers"])
    assert ("IMEI", "35533017092057") in ids
    assert ("PHONE", "+917500107305") in ids


def test_ipdr_prefix64_attaches_msisdn_to_host_sessions():
    t0 = datetime(2024, 11, 25, 13, 21, tzinfo=IST)
    trai = {
        "event_type": "IP_SESSION",
        "timestamp_start": t0, "timestamp_end": t0,
        "primary": ("PHONE", "+917500107305"),
        "own_identifiers": [
            ("PHONE", "+917500107305"),
            ("IP", "2409:40d2:132d:b3f8:8000:0000:0000:0000"),
        ],
        "attributes": {
            "public_ip": None,
            "private_ip": "2409:40d2:132d:b3f8:8000:0000:0000:0000",
        },
        "provenance": {"source_file": "trai.csv"},
    }
    host = {
        "event_type": "IP_SESSION",
        "timestamp_start": t0, "timestamp_end": t0,
        "primary": ("IP", "2409:40d2:132d:b3f8:cc81:c1fe:9ba7:e3cc"),
        "own_identifiers": [("IP", "2409:40d2:132d:b3f8:cc81:c1fe:9ba7:e3cc")],
        "attributes": {
            "public_ip": "2409:40d2:132d:b3f8:cc81:c1fe:9ba7:e3cc",
            "private_ip": None,
        },
        "provenance": {"source_file": "iprange.xlsx"},
    }
    n = norm.enrich_ipdr_prefix_phones([trai, host])
    assert n == 1
    assert host["primary"] == ("PHONE", "+917500107305")
    assert ("PHONE", "+917500107305") in host["own_identifiers"]


def test_cdr_imei_attaches_only_when_a_party_is_target():
    t0 = datetime(2024, 6, 1, 12, 0, tzinfo=IST)
    # Target matches A-party → IMEI becomes a merge key
    hit = norm._norm_cdr(
        {"entity_phone": "9876543210", "target_phone": "9876543210",
         "counterparty_phone": "9123456789", "timestamp_start": t0.isoformat(),
         "imei": "356938035643809"},
        {"source_file": "cdr.csv"}, "IST")
    assert ("IMEI", "356938035643809") in hit["own_identifiers"]

    # A-party is someone else → IMEI stays attributes-only (no merge)
    miss = norm._norm_cdr(
        {"entity_phone": "9123456789", "target_phone": "9876543210",
         "counterparty_phone": "9876543210", "timestamp_start": t0.isoformat(),
         "imei": "356938035643809"},
        {"source_file": "cdr.csv"}, "IST")
    assert all(t != "IMEI" for t, _ in miss["own_identifiers"])


def test_common_imei_link_merges_two_phones_on_same_handset(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Number", "Count", "9876543210", "9123456789"])
    ws.append(["356938035643809", "2", "Yes", "Yes"])
    wb.save(tmp_path / "CDR__x__Common_IMEI_Report.xlsx")

    events = [
        {"event_type": "CALL", "primary": ("PHONE", "+919876543210"),
         "own_identifiers": [("PHONE", "+919876543210")], "timestamp_start": None},
        {"event_type": "CALL", "primary": ("PHONE", "+919123456789"),
         "own_identifiers": [("PHONE", "+919123456789")], "timestamp_start": None},
    ]
    links = common_imei.load_common_imei_links(str(tmp_path), events=[])
    entities, n2e = er.resolve(events + links)
    assert n2e[("PHONE", "+919876543210")] == n2e[("PHONE", "+919123456789")]
